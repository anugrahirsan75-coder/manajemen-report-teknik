# -*- coding: utf-8 -*-
"""
Bahan presentasi direksi — Kendali Anggaran Teknik Cabang Ternate 2026.

Angka dihitung ULANG dari sumbernya, dengan rumus yang sama persis dengan
aplikasi Manajemen Report Teknik:
  · dokumen bertanda stokPersediaan TIDAK menggerus pagu (barang masuk gudang);
  · nilai baris = (harga SPBJ bila dokumennya sudah punya harga final, kalau
    tidak harga usulan) x jumlah;
  · jenis anggaran boleh ditimpa per BARIS, bukan hanya per dokumen;
  · satu baris yang menyebut beberapa kapal dibagi rata ke kapal-kapal itu.
Hasilnya sudah dicocokkan dengan layar aplikasi: pagu 3.401.910.619 dan
terpakai 3.119.939.030 — keduanya sama persis.
"""
import json, re, collections, datetime, os, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import DataBarRule

S = os.path.dirname(os.path.abspath(__file__))
AKAR = r"D:\ASDP\02. PROJEK\files\generator-swakelola"
KONTROL = r"D:\ASDP\01. ASDP TERNATE\2024 ASDP TERNATE\ASDP TERNATE\PERENCANAAN DOCKING\2026\KONTROL ANGGARAN TERNATE.xlsx"
KELUAR_DIR = os.path.join(AKAR, "output", "presentasi")
os.makedirs(KELUAR_DIR, exist_ok=True)
KELUAR = os.path.join(KELUAR_DIR, "Kendali Anggaran Teknik Ternate 2026 - Bahan Direksi.xlsx")

BULAN_NAMA = ["Januari", "Februari", "Maret", "April", "Mei", "Juni",
              "Juli", "Agustus", "September", "Oktober", "November", "Desember"]

# ── warna & gaya ──────────────────────────────────────────────────────────────
BIRU = "16357F"; BIRU_MUDA = "1CA3DD"; TOSCA = "14B8C4"
ABU = "F1F5F9"; ABU_GARIS = "CBD5E1"
MERAH = "DC2626"; MERAH_MUDA = "FEE2E2"
ORANYE = "F59E0B"; ORANYE_MUDA = "FEF3C7"
HIJAU = "16A34A"; HIJAU_MUDA = "DCFCE7"
PUTIH = "FFFFFF"

tepi = Side(style="thin", color=ABU_GARIS)
KOTAK = Border(left=tepi, right=tepi, top=tepi, bottom=tepi)
RP = '#,##0;[Red]-#,##0'
PERSEN = '0.0%'


def judul(ws, baris, teks, sub="", lebar=8):
    ws.merge_cells(start_row=baris, start_column=1, end_row=baris, end_column=lebar)
    s = ws.cell(baris, 1, teks)
    s.font = Font(name="Calibri", size=16, bold=True, color=BIRU)
    s.alignment = Alignment(vertical="center")
    ws.row_dimensions[baris].height = 24
    if sub:
        ws.merge_cells(start_row=baris + 1, start_column=1, end_row=baris + 1, end_column=lebar)
        t = ws.cell(baris + 1, 1, sub)
        t.font = Font(name="Calibri", size=10, color="475569")
    return baris + (3 if sub else 2)


def kepala_tabel(ws, baris, kolom, lebar_kolom=None):
    for i, teks in enumerate(kolom, start=1):
        c = ws.cell(baris, i, teks)
        c.font = Font(bold=True, color=PUTIH, size=10)
        c.fill = PatternFill("solid", fgColor=BIRU)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = KOTAK
    ws.row_dimensions[baris].height = 30
    if lebar_kolom:
        for i, w in enumerate(lebar_kolom, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return baris + 1


def sel(ws, r, c, nilai, fmt=None, tebal=False, isi=None, warna=None, rata=None, kotak=True):
    x = ws.cell(r, c, nilai)
    if fmt: x.number_format = fmt
    f = {"bold": tebal, "size": 10}
    if warna: f["color"] = warna
    x.font = Font(**f)
    if isi: x.fill = PatternFill("solid", fgColor=isi)
    if rata: x.alignment = Alignment(horizontal=rata, vertical="center")
    if kotak: x.border = KOTAK
    return x


def warna_status(rasio):
    """Melewati pagu = merah, 95-100% = waspada, sisanya aman — sama dengan lampu di aplikasi."""
    if rasio is None: return None, None
    if rasio > 1.0: return MERAH_MUDA, MERAH
    if rasio >= 0.95: return ORANYE_MUDA, "92400E"
    return HIJAU_MUDA, "166534"


# ══════════════════════════════════════════════════════════════════════════════
# 1. MUAT DATA
# ══════════════════════════════════════════════════════════════════════════════
pay = [x["payload"] for x in json.load(open(os.path.join(S, "db.json"), encoding="utf8"))]
RKA = json.load(open(os.path.join(AKAR, "src", "lib", "anggaran", "rka2026.json"), encoding="utf8"))
ang = [p for p in pay if p.get("kind") == "anggaran"][0]
ships = ([p for p in pay if p.get("kind") == "kapal"] or [{}])[0].get("ships") or []

maKey = lambda s: (re.search(r"\d{6,}", str(s or "")) or [None])[0] if re.search(r"\d{6,}", str(s or "")) else str(s or "").strip()
LABEL_MA = {
    "5010303001": "Bahan Pelumas", "5010403003": "Kapal Ro-Ro / Penyeberangan",
    "5010403009": "Akomodasi, Peralatan & Perlengkapan", "5010403100": "Permesinan & Kelistrikan",
    "5010302004": "Mobilisasi Docking", "5010302006": "Fumigasi",
    "5010318000": "Sertifikasi Docking", "5010103004": "Insentif Operasional",
    "1020604008": "Investasi Kapal Ro-Ro", "1020604009": "Investasi Akomodasi",
    "1020604010": "Investasi Permesinan",
}
nama_ma = lambda k: LABEL_MA.get(k, k or "(tanpa mata anggaran)")

KAPAL_URUT = ["KMP. TUNA", "KMP. MAMING", "KMP. PULAU SAGORI", "KMP. PORTLINK VIII", "KMP. LOMPA",
              "KMP. NGAFI", "KMP. KOLORAI", "KMP. BARONANG", "KMP. KERAPU II", "KMP. GORANGO",
              "KMP. BOBARA", "KMP. ARIWANGAN", "KMP. LEMA"]

def rapikan_kapal(nama):
    t = re.sub(r"\s+", " ", str(nama or "")).strip().upper()
    t = t.replace("KM.", "KMP.").replace("KMP ", "KMP. ")
    for k in KAPAL_URUT:
        if k.upper().replace("KMP. ", "") in t: return k
    return t or "(tanpa kapal)"


def pecah_kapal(teks):
    t = re.sub(r"\s+", " ", str(teks or "")).strip()
    if not t: return []
    hit = []
    up = t.upper()
    for k in KAPAL_URUT:
        inti = k.replace("KMP. ", "")
        if inti in up: hit.append(k)
    return hit or [rapikan_kapal(t)]


def jenis_dok(p):
    j = (p.get("jenisAnggaran") or "").lower()
    if j.startswith("dock"): return "docking"
    if j.startswith("lain"): return "lainnya"
    if j.startswith("rutin"): return "rutin"
    if p.get("programId"): return "lainnya"
    t = f"{p.get('kategoriRekap') or ''} {p.get('namaPengadaan') or ''}".lower()
    return "docking" if "docking" in t else "rutin"


def jenis_item(p, it):
    j = (it.get("jenisAnggaran") or "").lower()
    return j if j in ("rutin", "docking", "lainnya") else jenis_dok(p)


DOK = [p for p in pay if p.get("kind") in ("sppbj", "nonpr") and not p.get("stokPersediaan")]
DOK_STOK = [p for p in pay if p.get("kind") in ("sppbj", "nonpr") and p.get("stokPersediaan")]


def ma_dokumen(p):
    """Mata Anggaran dokumen. SPPBJ menyimpannya sebagai DAFTAR, Non PR PO sebagai
    TEKS — mengambil elemen [0] tanpa memeriksa jenisnya menghasilkan satu huruf
    ("5") dan memunculkan kolom mata anggaran palsu di rekap."""
    ma = p.get("mataAnggaran")
    if isinstance(ma, str): return ma
    if isinstance(ma, (list, tuple)) and ma: return ma[0]
    return ""


def baris_nilai(p):
    """hasilkan (jenis, ma, kapal, nilai) tiap baris pengadaan"""
    items = p.get("items") or []
    hf = any((it.get("hargaSpbj") or 0) > 0 for it in items)
    ma_default = ma_dokumen(p)
    for it in items:
        h = (it.get("hargaSpbj") or it.get("harga") or 0) if hf else (it.get("harga") or 0)
        v = h * (it.get("jumlah") or 0)
        if not v: continue
        ma = maKey((it.get("mataAnggaran") or "").strip() or ma_default)
        ks = pecah_kapal(it.get("kapal"))
        if not ks: ks = ["(tanpa kapal)"]
        for k in ks:
            yield jenis_item(p, it), ma, k, v / len(ks)


# ── realisasi ────────────────────────────────────────────────────────────────
real_bulan = collections.defaultdict(float)                      # rutin per bulan
real_ma = collections.defaultdict(float)                         # rutin per MA
real_kapal = collections.defaultdict(float)                      # rutin per kapal
real_kapal_ma = collections.defaultdict(float)                   # rutin (kapal, MA)
real_kapal_bulan = collections.defaultdict(float)                # rutin (kapal, bulan)
dock_kapal = collections.defaultdict(float)
dock_kapal_ma = collections.defaultdict(float)
lain_kapal = collections.defaultdict(float)
lain_program = collections.defaultdict(float)
dokumen = []                                                     # rincian untuk sheet data

for p in DOK:
    bulan = (p.get("tanggal") or "")[:7]
    nilai_dok = collections.defaultdict(float)
    for jenis, ma, kapal, v in baris_nilai(p):
        if jenis == "rutin":
            real_bulan[bulan] += v; real_ma[ma] += v
            real_kapal[kapal] += v; real_kapal_ma[(kapal, ma)] += v
            real_kapal_bulan[(kapal, bulan)] += v
        elif jenis == "docking":
            dock_kapal[kapal] += v; dock_kapal_ma[(kapal, ma)] += v
        else:
            lain_kapal[kapal] += v
            lain_program[p.get("programId") or "(tanpa program)"] += v
        nilai_dok[(jenis, ma)] += v
    for (jenis, ma), v in nilai_dok.items():
        dokumen.append({
            "tanggal": p.get("tanggal") or "", "jenis": jenis, "ma": ma,
            "nama": p.get("namaPengadaan") or "(tanpa nama)",
            "nilai": v, "status": p.get("status") or "", "sumber": p.get("kind"),
        })

# ── pagu (Persetujuan Pusat) ─────────────────────────────────────────────────
pagu_bulan = collections.defaultdict(float)
pagu_bulan_ma = collections.defaultdict(float)
for blok in ang.get("plafon") or []:
    b = blok.get("bulan")
    for r in blok.get("rows") or []:
        v = (r.get("nilai") or 0) + (r.get("addendum") or 0)
        pagu_bulan[b] += v
        pagu_bulan_ma[(b, maKey(r.get("ma")))] += v

pagu_docking = collections.defaultdict(float)
for blok in ang.get("docking") or []:
    kapal = rapikan_kapal(blok.get("kapal") or "")
    for r in blok.get("rows") or []:
        pagu_docking[kapal] += (r.get("nilai") or 0) + (r.get("addendum") or 0)

# ── RKA rutin per kapal (mengikuti aturan aplikasi) ──────────────────────────
MA_DOCKING_DI_RUTIN = {"5010403003", "5010302004", "5010302006", "5010318000"}

def bulan_docking(pos):
    arr = pos.get("5010403003") or []
    for i, v in enumerate(arr):
        if v: return i + 1
    return 0

rka_kapal_bulan = collections.defaultdict(float)
rka_kapal_ma = collections.defaultdict(float)
for nama_kapal, pos in (RKA.get("rutin") or {}).items():
    k = rapikan_kapal(nama_kapal)
    bd = bulan_docking(pos)
    for ma, arr in pos.items():
        if ma in MA_DOCKING_DI_RUTIN: continue
        for i, v in enumerate(arr or []):
            if not v or (i + 1) == bd: continue      # bulan docking bukan anggaran rutin
            rka_kapal_bulan[(k, f"2026-{i+1:02d}")] += v
            rka_kapal_ma[(k, ma)] += v

# docking: tiap Mata Anggaran punya rka/tambahan/persetujuan/release dari berkas cabang
rka_docking_kapal = collections.defaultdict(float)
dock_pos = collections.defaultdict(lambda: collections.defaultdict(float))   # kapal -> pos -> nilai
for nama_kapal, pos in (RKA.get("docking") or {}).items():
    k = rapikan_kapal(nama_kapal)
    for ma, isi in (pos or {}).items():
        if not isinstance(isi, dict): continue
        rka_docking_kapal[k] += isi.get("rka") or 0
        for kunci in ("rka", "tambahan", "persetujuan", "release", "realisasi"):
            dock_pos[k][kunci] += isi.get(kunci) or 0

# ── data kapal (GT & tahun) ──────────────────────────────────────────────────
info_kapal = {}
for s in ships:
    k = rapikan_kapal(s.get("nama"))
    dim = s.get("dimension") or {}
    gen = s.get("general") or {}
    gt = None
    for kunci in ("gt", "grt", "isiKotor", "tonaseKotor"):
        v = dim.get(kunci) or gen.get(kunci)
        if v:
            m = re.search(r"[\d.,]+", str(v))
            if m: gt = float(m.group(0).replace(".", "").replace(",", "."))
            break
    tahun = None
    m = re.search(r"(19|20)\d{2}", str(gen.get("tahun") or ""))
    if m: tahun = int(m.group(0))
    info_kapal[k] = {"gt": gt, "tahun": tahun, "lintasan": gen.get("lintasan") or ""}

# ── pergeseran anggaran ──────────────────────────────────────────────────────
geser = []
try:
    wbk = openpyxl.load_workbook(KONTROL, data_only=True)
    wsg = wbk["HISTORY GESER"]
    for r in range(2, wsg.max_row + 1):
        if not wsg.cell(r, 2).value: continue
        geser.append({
            "jenisAsal": wsg.cell(r, 2).value, "ma": wsg.cell(r, 3).value,
            "kapalAsal": rapikan_kapal(wsg.cell(r, 4).value), "bulanAsal": wsg.cell(r, 6).value,
            "maTujuan": wsg.cell(r, 7).value, "kapalTujuan": rapikan_kapal(wsg.cell(r, 8).value),
            "jenisTujuan": wsg.cell(r, 10).value, "nominal": wsg.cell(r, 11).value or 0,
        })
except Exception as e:
    print("HISTORY GESER tidak terbaca:", e)

BULAN_ADA = sorted(set(list(pagu_bulan) + [b for b in real_bulan if b]))
BULAN_ADA = [b for b in BULAN_ADA if b.startswith("2026")]
TOT_PAGU = sum(pagu_bulan.values())
TOT_REAL = sum(real_bulan[b] for b in BULAN_ADA)
print(f"pagu {TOT_PAGU:,.0f} | realisasi {TOT_REAL:,.0f} | sisa {TOT_PAGU-TOT_REAL:,.0f}")

# ══════════════════════════════════════════════════════════════════════════════
# 2. BANGUN WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ─────────────────────────── SHEET 1: RINGKASAN ──────────────────────────────
ws = wb.active
ws.title = "1. Ringkasan"
ws.sheet_view.showGridLines = False
for i, w in enumerate([3, 34, 20, 20, 20, 20, 20, 22, 20], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.merge_cells("B2:I3")
t = ws.cell(2, 2, "KENDALI ANGGARAN TEKNIK — CABANG TERNATE 2026")
t.font = Font(size=20, bold=True, color=BIRU)
ws.merge_cells("B4:I4")
ws.cell(4, 2, f"Bahan rapat Direksi · data per {datetime.date.today().strftime('%d %B %Y')} · "
              f"sumber: SPPBJ & Non PR PO aplikasi Manajemen Report Teknik + KONTROL ANGGARAN TERNATE.xlsx").font = Font(size=10, color="475569")

kpi = [
    ("PAGU RUTIN (Jan–Ags)", TOT_PAGU, RP, BIRU),
    ("REALISASI RUTIN", TOT_REAL, RP, MERAH if TOT_REAL > TOT_PAGU else BIRU_MUDA),
    ("SISA", TOT_PAGU - TOT_REAL, RP, HIJAU if TOT_PAGU - TOT_REAL > 0 else MERAH),
    ("SERAPAN", TOT_REAL / TOT_PAGU if TOT_PAGU else 0, PERSEN, ORANYE),
]
r0 = 6
for i, (lab, val, fmt, warna) in enumerate(kpi):
    c = 2 + i * 2
    ws.merge_cells(start_row=r0, start_column=c, end_row=r0, end_column=c + 1)
    x = ws.cell(r0, c, lab); x.font = Font(size=9, bold=True, color="64748B")
    ws.merge_cells(start_row=r0 + 1, start_column=c, end_row=r0 + 1, end_column=c + 1)
    y = ws.cell(r0 + 1, c, val); y.font = Font(size=18, bold=True, color=warna); y.number_format = fmt
ws.row_dimensions[r0 + 1].height = 26

# tabel bulanan ringkas
r = r0 + 4
ws.cell(r, 2, "PAGU VS REALISASI PER BULAN").font = Font(size=12, bold=True, color=BIRU)
r += 1
head = kepala_tabel(ws, r, ["", "Bulan", "Pagu (Persetujuan Pusat)", "Realisasi", "Serapan", "Selisih", "Status"], None)
awal_data = head
for b in BULAN_ADA:
    pg = pagu_bulan.get(b, 0); rl = real_bulan.get(b, 0)
    ras = (rl / pg) if pg else None
    isi, wr = warna_status(ras)
    sel(ws, head, 2, BULAN_NAMA[int(b[5:7]) - 1], tebal=True)
    sel(ws, head, 3, pg, RP)
    sel(ws, head, 4, rl, RP)
    sel(ws, head, 5, ras if ras is not None else "", PERSEN, isi=isi, warna=wr, tebal=True, rata="center")
    sel(ws, head, 6, pg - rl, RP, warna=MERAH if rl > pg else None)
    sel(ws, head, 7, "MELEWATI PAGU" if ras and ras > 1.0 else ("Waspada" if ras and ras >= 0.95 else ("Aman" if pg else "Belum ada pagu")),
        isi=isi, warna=wr, rata="center")
    head += 1
sel(ws, head, 2, "TOTAL", tebal=True, isi=ABU)
sel(ws, head, 3, TOT_PAGU, RP, tebal=True, isi=ABU)
sel(ws, head, 4, TOT_REAL, RP, tebal=True, isi=ABU)
sel(ws, head, 5, TOT_REAL / TOT_PAGU if TOT_PAGU else 0, PERSEN, tebal=True, isi=ABU, rata="center")
sel(ws, head, 6, TOT_PAGU - TOT_REAL, RP, tebal=True, isi=ABU)
sel(ws, head, 7, "", isi=ABU)
akhir_data = head

# grafik pagu vs realisasi + garis serapan
ch = BarChart(); ch.type = "col"; ch.style = 10
ch.title = "Pagu vs Realisasi Rutin per Bulan"
data = Reference(ws, min_col=3, max_col=4, min_row=awal_data - 1, max_row=akhir_data - 1)
kat = Reference(ws, min_col=2, min_row=awal_data, max_row=akhir_data - 1)
ch.add_data(data, titles_from_data=True); ch.set_categories(kat)
ch.y_axis.numFmt = '#,##0'
ch.height, ch.width = 9, 24
ch.gapWidth = 60
ch.legend.position = "b"
ch.legend.overlay = False
ch.x_axis.delete = False; ch.y_axis.delete = False
gs = LineChart()
gsdata = Reference(ws, min_col=5, min_row=awal_data - 1, max_row=akhir_data - 1)
gs.add_data(gsdata, titles_from_data=True)
gs.y_axis.axId = 200; gs.y_axis.numFmt = '0%'; gs.y_axis.delete = False
gs.y_axis.crosses = "max"          # sumbu persen di sisi kanan
ch += gs
ch.x_axis.delete = False; ch.y_axis.delete = False
ch.legend.position = "b"
ch.legend.overlay = False
ws.add_chart(ch, f"B{akhir_data + 2}")

# ══════════════════════════════════════════════════════════════════════════════
# 3. ANALISIS — dihitung dulu, supaya sheet & rekomendasi memakai angka yang sama
# ══════════════════════════════════════════════════════════════════════════════
BULAN_JALAN = [b for b in BULAN_ADA if real_bulan.get(b, 0) > 0]          # bulan yang sudah ada belanja
BULAN_OVER = [(b, pagu_bulan.get(b, 0), real_bulan.get(b, 0)) for b in BULAN_ADA
              if pagu_bulan.get(b, 0) and real_bulan.get(b, 0) / pagu_bulan[b] > 1.0]
KELEBIHAN = sum(r - p for _, p, r in BULAN_OVER)

# per kapal: RKA vs realisasi rutin
kapal_baris = []
for k in sorted(set(list(real_kapal) + [x[0] for x in rka_kapal_bulan])):
    if k == "(tanpa kapal)": continue
    rka = sum(v for (kk, b), v in rka_kapal_bulan.items() if kk == k and b in BULAN_ADA)
    real = real_kapal.get(k, 0)
    info = info_kapal.get(k, {})
    gt = info.get("gt") or 0
    umur = (2026 - info["tahun"]) if info.get("tahun") else None
    kapal_baris.append({
        "kapal": k, "rka": rka, "real": real,
        "rasio": (real / rka) if rka else None,
        "selisih": real - rka,
        "gt": gt, "perGt": (real / gt) if gt else None,
        "umur": umur,
        "docking": dock_kapal.get(k, 0),
        "lainnya": lain_kapal.get(k, 0),
    })
kapal_baris.sort(key=lambda x: (x["rasio"] is None, -(x["rasio"] or 0)))
BOROS = [x for x in kapal_baris if x["rasio"] and x["rasio"] > 1][:5]

# penyumbang terbesar tiap kapal
def penyumbang(k, n=3):
    pos = [(ma, v) for (kk, ma), v in real_kapal_ma.items() if kk == k and v > 0]
    pos.sort(key=lambda x: -x[1])
    tot = sum(v for _, v in pos) or 1
    return [(nama_ma(ma), v, v / tot) for ma, v in pos[:n]]

# pengadaan rutin terbesar (tanpa vendor)
rutin_dok = [d for d in dokumen if d["jenis"] == "rutin"]
rutin_dok.sort(key=lambda d: -d["nilai"])

# proyeksi sampai Desember
bulan_terpakai = len(BULAN_JALAN)
rata_real = TOT_REAL / bulan_terpakai if bulan_terpakai else 0
sisa_bulan = 12 - bulan_terpakai
proyeksi_akhir = TOT_REAL + rata_real * sisa_bulan
pagu_setahun_perkiraan = TOT_PAGU / len(BULAN_ADA) * 12 if BULAN_ADA else 0

# RKA 2027: kapal yang RKA-nya konsisten kekecilan / kebesaran
usul_rka = sorted([x for x in kapal_baris if x["rka"] > 0],
                  key=lambda x: -(x["real"] - x["rka"]))

# pergeseran anggaran
geser_total = sum(g["nominal"] or 0 for g in geser)
geser_per_tujuan = collections.Counter()
for g in geser:
    geser_per_tujuan[str(g.get("jenisTujuan") or "-")] += g["nominal"] or 0

nilai_stok = 0
for p in DOK_STOK:
    items = p.get("items") or []
    hf = any((it.get("hargaSpbj") or 0) > 0 for it in items)
    for it in items:
        h = (it.get("hargaSpbj") or it.get("harga") or 0) if hf else (it.get("harga") or 0)
        nilai_stok += h * (it.get("jumlah") or 0)


# ══════════════════════════════════════════════════════════════════════════════
# 4. LANJUTAN SHEET 1 — REKOMENDASI & PROYEKSI
# ══════════════════════════════════════════════════════════════════════════════
r = akhir_data + 20
ws.cell(r, 2, "TIGA HAL YANG DIMINTAKAN KEPUTUSAN").font = Font(size=12, bold=True, color=BIRU)
r += 1
minta = [
    ("1. Tambahan pagu / realokasi",
     f"{len(BULAN_OVER)} bulan melewati pagu dengan kelebihan {KELEBIHAN:,.0f}. "
     f"Sisa pagu {TOT_PAGU-TOT_REAL:,.0f} untuk {12-len(BULAN_ADA)+1} bulan sisa tahun."),
    ("2. Evaluasi kapal berbiaya tinggi",
     (f"{BOROS[0]['kapal']} menyerap {BOROS[0]['rasio']*100:.0f}% dari RKA rutinnya "
      f"({BOROS[0]['real']:,.0f} vs {BOROS[0]['rka']:,.0f})." ) if BOROS else "Tidak ada kapal yang melampaui RKA."),
    ("3. Dasar penyusunan RKA 2027",
     f"{sum(1 for x in kapal_baris if x['rasio'] and x['rasio']>1)} dari {len(kapal_baris)} kapal "
     f"realisasinya melampaui RKA — angka RKA perlu disesuaikan, bukan ditekan."),
]
for tajuk, isi in minta:
    sel(ws, r, 2, tajuk, tebal=True, isi=ABU, kotak=True)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
    sel(ws, r, 3, isi, isi=ABU)
    r += 1

r += 2
ws.cell(r, 2, "REKOMENDASI (setiap butir menyebut angka dasarnya)").font = Font(size=12, bold=True, color=BIRU)
r += 1
rek = []
if BULAN_OVER:
    rinci = ", ".join(f"{BULAN_NAMA[int(b[5:7])-1]} {rl/pg*100:.0f}%" for b, pg, rl in BULAN_OVER)
    rek.append((f"Ajukan tambahan pagu / realokasi sebesar {KELEBIHAN:,.0f}",
                f"Bulan yang melewati pagu: {rinci}. Kelebihan ini sudah terjadi, bukan rencana."))
if proyeksi_akhir > pagu_setahun_perkiraan and pagu_setahun_perkiraan:
    rek.append((f"Siapkan tambahan ±{proyeksi_akhir-pagu_setahun_perkiraan:,.0f} sampai Desember",
                f"Rata-rata belanja {rata_real:,.0f}/bulan selama {bulan_terpakai} bulan. "
                f"Bila berlanjut, realisasi setahun ±{proyeksi_akhir:,.0f} sedangkan pagu setahun ±{pagu_setahun_perkiraan:,.0f}."))
for x in BOROS[:3]:
    pos = penyumbang(x["kapal"])
    sebab = "; ".join(f"{n} {v/1e6:,.0f} jt ({s*100:.0f}%)" for n, v, s in pos)
    rek.append((f"Audit teknis {x['kapal']} — serapan {x['rasio']*100:.0f}% dari RKA",
                f"Realisasi {x['real']:,.0f} vs RKA {x['rka']:,.0f} (lebih {x['selisih']:,.0f}). Penyumbang: {sebab}."))
if usul_rka:
    naik = usul_rka[0]
    rek.append((f"Naikkan RKA 2027 {naik['kapal']} minimal {naik['real']-naik['rka']:,.0f}",
                f"Selama {bulan_terpakai} bulan 2026 realisasinya {naik['real']:,.0f} sedangkan RKA {naik['rka']:,.0f}. "
                f"RKA yang kekecilan memaksa pergeseran anggaran berulang."))
if geser_total:
    rek.append((f"Tertibkan pergeseran anggaran — {len(geser)} kali, total {geser_total:,.0f}",
                "Pergeseran sebanyak ini menandakan rencana awal tidak sesuai kebutuhan lapangan; "
                "lihat sheet 7 untuk asal dan tujuannya."))
if nilai_stok:
    rek.append((f"Catat pemakaian stok persediaan {nilai_stok:,.0f}",
                "Nilai ini TIDAK menggerus pagu saat dibeli, tetapi menjadi beban saat dipakai. "
                "Tanpa pencatatan pemakaian, biaya kapal terlihat lebih murah dari kenyataan."))

for i, (tajuk, dasar) in enumerate(rek, start=1):
    sel(ws, r, 2, f"R{i}", tebal=True, isi=BIRU, warna=PUTIH, rata="center")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
    sel(ws, r, 3, tajuk, tebal=True)
    r += 1
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
    sel(ws, r, 3, dasar, warna="475569")
    sel(ws, r, 2, "", kotak=True)
    r += 1

# tabel proyeksi kecil + grafik kumulatif
r += 2
ws.cell(r, 2, "PROYEKSI SAMPAI DESEMBER (garis putus = perkiraan)").font = Font(size=12, bold=True, color=BIRU)
r += 1
pr_awal = kepala_tabel(ws, r, ["", "Bulan", "Realisasi kumulatif", "Pagu kumulatif", "Proyeksi kumulatif"], None)
kum_r = kum_p = 0
baris_proyeksi = pr_awal
pagu_rata = TOT_PAGU / len(BULAN_ADA) if BULAN_ADA else 0
for i in range(12):
    b = f"2026-{i+1:02d}"
    nyata = real_bulan.get(b, 0)
    pg = pagu_bulan.get(b, pagu_rata)
    kum_p += pg
    sel(ws, baris_proyeksi, 2, BULAN_NAMA[i], tebal=True)
    if b in BULAN_JALAN:
        kum_r += nyata
        sel(ws, baris_proyeksi, 3, kum_r, RP)
        sel(ws, baris_proyeksi, 5, kum_r, RP)
    else:
        sel(ws, baris_proyeksi, 3, None, RP)
        sel(ws, baris_proyeksi, 5, kum_r + rata_real * (i + 1 - bulan_terpakai), RP, warna="94A3B8")
    sel(ws, baris_proyeksi, 4, kum_p, RP)
    baris_proyeksi += 1

grafik = LineChart(); grafik.title = "Kumulatif: realisasi vs pagu vs proyeksi"
grafik.style = 12; grafik.height, grafik.width = 9, 22
d1 = Reference(ws, min_col=3, max_col=5, min_row=pr_awal - 1, max_row=baris_proyeksi - 1)
k1 = Reference(ws, min_col=2, min_row=pr_awal, max_row=baris_proyeksi - 1)
grafik.add_data(d1, titles_from_data=True); grafik.set_categories(k1)
grafik.y_axis.numFmt = '#,##0'
grafik.x_axis.delete = False; grafik.y_axis.delete = False
grafik.legend.position = "b"
grafik.legend.overlay = False
ws.add_chart(grafik, f"F{r}")

ws.freeze_panes = "A6"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — RUTIN PER MATA ANGGARAN & BULAN
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2. Rutin per Bulan")
ws2.sheet_view.showGridLines = False
r = judul(ws2, 2, "RUTIN — PAGU DAN REALISASI PER MATA ANGGARAN",
          "Pagu = Persetujuan Pusat per bulan. Realisasi = SPPBJ + Non PR PO berjenis Rutin, "
          "tidak termasuk barang yang masuk persediaan.", lebar=9)

# kolom mata anggaran: hanya yang punya pagu ATAU realisasi. Pos yang nol
# sepanjang tahun cuma menambah kolom kosong dan memperkecil angka penting.
ma_dipakai = sorted(
    {k for (_, k), v in pagu_bulan_ma.items() if v} | {k for k, v in real_ma.items() if k and v},
    key=lambda k: -(real_ma.get(k, 0)))
kol = ["", "Bulan"] + [nama_ma(k) for k in ma_dipakai] + ["Total realisasi", "Total pagu", "Serapan"]
lebar = [3, 14] + [20] * len(ma_dipakai) + [18, 18, 11]
h = kepala_tabel(ws2, r, kol, lebar)
awal2 = h
for b in BULAN_ADA:
    sel(ws2, h, 2, BULAN_NAMA[int(b[5:7]) - 1], tebal=True)
    tot_r = 0
    for i, k in enumerate(ma_dipakai):
        # realisasi per MA per bulan
        v = sum(d["nilai"] for d in rutin_dok if d["ma"] == k and d["tanggal"][:7] == b)
        pg = pagu_bulan_ma.get((b, k), 0)
        ras = (v / pg) if pg else None
        isi, wr = warna_status(ras)
        sel(ws2, h, 3 + i, v, RP, isi=isi, warna=wr)
        tot_r += v
    sel(ws2, h, 3 + len(ma_dipakai), tot_r, RP, tebal=True)
    sel(ws2, h, 4 + len(ma_dipakai), pagu_bulan.get(b, 0), RP)
    ras = tot_r / pagu_bulan[b] if pagu_bulan.get(b) else None
    isi, wr = warna_status(ras)
    sel(ws2, h, 5 + len(ma_dipakai), ras if ras is not None else "", PERSEN, isi=isi, warna=wr, tebal=True, rata="center")
    h += 1
sel(ws2, h, 2, "TOTAL", tebal=True, isi=ABU)
for i, k in enumerate(ma_dipakai):
    sel(ws2, h, 3 + i, real_ma.get(k, 0), RP, tebal=True, isi=ABU)
sel(ws2, h, 3 + len(ma_dipakai), TOT_REAL, RP, tebal=True, isi=ABU)
sel(ws2, h, 4 + len(ma_dipakai), TOT_PAGU, RP, tebal=True, isi=ABU)
sel(ws2, h, 5 + len(ma_dipakai), TOT_REAL / TOT_PAGU, PERSEN, tebal=True, isi=ABU, rata="center")

c2 = BarChart(); c2.type = "col"; c2.grouping = "stacked"; c2.overlap = 100
c2.title = "Belanja rutin per bulan menurut Mata Anggaran"
c2.height, c2.width = 10, 26
d2 = Reference(ws2, min_col=3, max_col=2 + len(ma_dipakai), min_row=awal2 - 1, max_row=h - 1)
k2 = Reference(ws2, min_col=2, min_row=awal2, max_row=h - 1)
c2.add_data(d2, titles_from_data=True); c2.set_categories(k2)
c2.y_axis.numFmt = '#,##0'
c2.x_axis.delete = False; c2.y_axis.delete = False
c2.legend.position = "b"
c2.legend.overlay = False
ws2.add_chart(c2, f"B{h + 3}")
ws2.freeze_panes = ws2.cell(awal2, 3)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — PER KAPAL (RKA vs realisasi)
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3. Per Kapal")
ws3.sheet_view.showGridLines = False
r = judul(ws3, 2, "PERBANDINGAN TERHADAP RKA MASING-MASING KAPAL",
          "Boros diukur dari RKA kapal itu sendiri, bukan dibanding kapal lain — kapal besar memang "
          "wajar lebih mahal. Kolom Rp/GT disediakan sebagai pembanding kedua.", lebar=10)
kol3 = ["", "Kapal", "RKA rutin (bulan berjalan)", "Realisasi rutin", "Serapan thd RKA", "Selisih",
        "GT", "Rp per GT", "Umur (th)", "Realisasi docking", "Realisasi lainnya"]
h = kepala_tabel(ws3, r, kol3, [3, 22, 20, 18, 13, 18, 9, 14, 10, 18, 16])
awal3 = h
for x in kapal_baris:
    isi, wr = warna_status(x["rasio"])
    sel(ws3, h, 2, x["kapal"], tebal=True)
    sel(ws3, h, 3, x["rka"], RP)
    sel(ws3, h, 4, x["real"], RP)
    sel(ws3, h, 5, x["rasio"] if x["rasio"] is not None else "", PERSEN, isi=isi, warna=wr, tebal=True, rata="center")
    sel(ws3, h, 6, x["selisih"], RP, warna=MERAH if x["selisih"] > 0 else None)
    sel(ws3, h, 7, x["gt"] or "", '#,##0')
    sel(ws3, h, 8, x["perGt"] or "", RP)
    sel(ws3, h, 9, x["umur"] or "", '0', rata="center")
    sel(ws3, h, 10, x["docking"], RP)
    sel(ws3, h, 11, x["lainnya"], RP)
    h += 1
sel(ws3, h, 2, "TOTAL", tebal=True, isi=ABU)
for kolom, kunci in ((3, "rka"), (4, "real"), (6, "selisih"), (10, "docking"), (11, "lainnya")):
    sel(ws3, h, kolom, sum(x[kunci] for x in kapal_baris), RP, tebal=True, isi=ABU)
for kolom in (5, 7, 8, 9):
    sel(ws3, h, kolom, "", isi=ABU)

c3 = BarChart(); c3.type = "bar"; c3.title = "RKA vs Realisasi rutin per kapal"
c3.height, c3.width = 12, 22
d3 = Reference(ws3, min_col=3, max_col=4, min_row=awal3 - 1, max_row=h - 1)
k3 = Reference(ws3, min_col=2, min_row=awal3, max_row=h - 1)
c3.add_data(d3, titles_from_data=True); c3.set_categories(k3)
c3.x_axis.numFmt = '#,##0'
c3.x_axis.delete = False; c3.y_axis.delete = False
c3.legend.position = "b"
c3.legend.overlay = False
ws3.add_chart(c3, f"B{h + 3}")

c3b = BarChart(); c3b.type = "bar"; c3b.title = "Serapan terhadap RKA (100% = tepat rencana)"
c3b.height, c3b.width = 12, 18
d3b = Reference(ws3, min_col=5, min_row=awal3 - 1, max_row=h - 1)
c3b.add_data(d3b, titles_from_data=True); c3b.set_categories(k3)
c3b.x_axis.numFmt = '0%'
c3b.x_axis.delete = False; c3b.y_axis.delete = False
c3b.legend.position = "b"
c3b.legend.overlay = False
ws3.add_chart(c3b, f"H{h + 3}")
ws3.freeze_panes = ws3.cell(awal3, 3)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — KENAPA BOROS
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4. Kenapa Boros")
ws4.sheet_view.showGridLines = False
r = judul(ws4, 2, "PEMBEDAHAN: DI POS MANA BIAYANYA JATUH",
          "Matriks kapal x Mata Anggaran untuk belanja rutin, disusul daftar pengadaan terbesar. "
          "Nama penyedia sengaja tidak ditampilkan.", lebar=9)
kol4 = ["", "Kapal"] + [nama_ma(k) for k in ma_dipakai] + ["Total", "Pos terbesar", "Porsi pos terbesar"]
h = kepala_tabel(ws4, r, kol4, [3, 22] + [19] * len(ma_dipakai) + [18, 34, 13])
awal4 = h
for x in kapal_baris:
    k = x["kapal"]
    sel(ws4, h, 2, k, tebal=True)
    nilai_pos = []
    for i, ma in enumerate(ma_dipakai):
        v = real_kapal_ma.get((k, ma), 0)
        nilai_pos.append((ma, v))
        sel(ws4, h, 3 + i, v, RP)
    tot = sum(v for _, v in nilai_pos)
    sel(ws4, h, 3 + len(ma_dipakai), tot, RP, tebal=True)
    if tot:
        ma_top, v_top = max(nilai_pos, key=lambda t: t[1])
        sel(ws4, h, 4 + len(ma_dipakai), nama_ma(ma_top))
        sel(ws4, h, 5 + len(ma_dipakai), v_top / tot, PERSEN, rata="center",
            isi=MERAH_MUDA if v_top / tot > 0.6 else None)
    else:
        sel(ws4, h, 4 + len(ma_dipakai), "-")
        sel(ws4, h, 5 + len(ma_dipakai), "")
    h += 1

c4 = BarChart(); c4.type = "col"; c4.grouping = "stacked"; c4.overlap = 100
c4.title = "Susunan biaya rutin tiap kapal"
c4.height, c4.width = 11, 26
d4 = Reference(ws4, min_col=3, max_col=2 + len(ma_dipakai), min_row=awal4 - 1, max_row=h - 1)
k4 = Reference(ws4, min_col=2, min_row=awal4, max_row=h - 1)
c4.add_data(d4, titles_from_data=True); c4.set_categories(k4)
c4.y_axis.numFmt = '#,##0'
c4.x_axis.delete = False; c4.y_axis.delete = False
c4.legend.position = "b"
c4.legend.overlay = False
ws4.add_chart(c4, f"B{h + 2}")

r5 = h + 26
ws4.cell(r5, 2, "20 PENGADAAN RUTIN TERBESAR").font = Font(size=12, bold=True, color=BIRU)
r5 += 1
h2 = kepala_tabel(ws4, r5, ["", "Tanggal", "Nama pengadaan", "Mata Anggaran", "Nilai", "Sumber", "Status"], None)
for d in rutin_dok[:20]:
    sel(ws4, h2, 2, d["tanggal"])
    sel(ws4, h2, 3, d["nama"][:70])
    sel(ws4, h2, 4, nama_ma(d["ma"]))
    sel(ws4, h2, 5, d["nilai"], RP, tebal=True)
    sel(ws4, h2, 6, "SPPBJ" if d["sumber"] == "sppbj" else "Non PR PO")
    sel(ws4, h2, 7, d["status"])
    h2 += 1
ws4.freeze_panes = ws4.cell(awal4, 3)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — DOCKING
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("5. Docking")
ws5.sheet_view.showGridLines = False
r = judul(ws5, 2, "DOCKING — RENCANA, PERSETUJUAN, DAN REALISASI",
          "RKA/persetujuan/release berasal dari berkas KONTROL ANGGARAN TERNATE; realisasi dihitung "
          "dari SPPBJ & Non PR PO berjenis Docking pada aplikasi.", lebar=9)
kol5 = ["", "Kapal", "RKA docking", "Anggaran tambahan", "Persetujuan", "Release",
        "Realisasi (aplikasi)", "Dasar pembanding", "Serapan", "Selisih"]
h = kepala_tabel(ws5, r, kol5, [3, 22, 18, 18, 18, 18, 20, 17, 11, 18])
awal5 = h
kapal_dock = sorted(set(list(dock_pos) + list(dock_kapal)), key=lambda k: -(dock_kapal.get(k, 0)))
for k in kapal_dock:
    if k == "(tanpa kapal)": continue
    pos = dock_pos.get(k, {})
    rka_d = pos.get("rka", 0)
    persetujuan = pos.get("persetujuan", 0) or pagu_docking.get(k, 0)
    real = dock_kapal.get(k, 0)
    # Kolom persetujuan di berkas cabang belum terisi penuh untuk sebagian kapal.
    # Membandingkan realisasi terhadap angka yang baru sebagian membuat serapan
    # tampak ribuan persen — menyesatkan di forum. Jadi dasar pembandingnya
    # dipilih terang-terangan, dan disebutkan yang mana yang dipakai.
    if persetujuan and rka_d and persetujuan >= 0.5 * rka_d:
        dasar, label_dasar = persetujuan, "Persetujuan"
    elif rka_d:
        dasar, label_dasar = rka_d, "RKA (persetujuan belum lengkap)"
    else:
        dasar, label_dasar = persetujuan, "Persetujuan"
    ras = (real / dasar) if dasar else None
    isi, wr = warna_status(ras)
    sel(ws5, h, 2, k, tebal=True)
    sel(ws5, h, 3, rka_d, RP)
    sel(ws5, h, 4, pos.get("tambahan", 0), RP)
    sel(ws5, h, 5, persetujuan, RP)
    sel(ws5, h, 6, pos.get("release", 0), RP)
    sel(ws5, h, 7, real, RP, tebal=True)
    sel(ws5, h, 8, label_dasar, warna="475569" if "belum" in label_dasar else None)
    sel(ws5, h, 9, ras if ras is not None else "", PERSEN, isi=isi, warna=wr, rata="center", tebal=True)
    sel(ws5, h, 10, dasar - real, RP, warna=MERAH if dasar and real > dasar else None)
    h += 1

c5 = BarChart(); c5.type = "col"; c5.title = "Docking: RKA vs realisasi per kapal"
c5.height, c5.width = 10, 24
d5 = Reference(ws5, min_col=3, max_col=3, min_row=awal5 - 1, max_row=h - 1)
k5 = Reference(ws5, min_col=2, min_row=awal5, max_row=h - 1)
c5.add_data(d5, titles_from_data=True)
d5b = Reference(ws5, min_col=7, max_col=7, min_row=awal5 - 1, max_row=h - 1)
c5.add_data(d5b, titles_from_data=True)
c5.set_categories(k5)
c5.y_axis.numFmt = '#,##0'
c5.x_axis.delete = False; c5.y_axis.delete = False
c5.legend.position = "b"
c5.legend.overlay = False
ws5.add_chart(c5, f"B{h + 2}")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — PERGESERAN ANGGARAN
# ══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("6. Pergeseran")
ws6.sheet_view.showGridLines = False
r = judul(ws6, 2, "PERGESERAN ANGGARAN — SEBAB BULAN TERTENTU MELONJAK",
          f"{len(geser)} kali pergeseran, total {geser_total:,.0f}. Sumber: sheet HISTORY GESER "
          "pada berkas KONTROL ANGGARAN TERNATE.", lebar=9)
kol6 = ["", "MA asal", "Kapal asal", "Bulan asal", "MA tujuan", "Kapal tujuan", "Jenis tujuan", "Nominal"]
h = kepala_tabel(ws6, r, kol6, [3, 20, 20, 13, 20, 20, 14, 18])
awal6 = h
for g in sorted(geser, key=lambda g: -(g["nominal"] or 0)):
    sel(ws6, h, 2, str(g.get("ma") or ""))
    sel(ws6, h, 3, g.get("kapalAsal") or "")
    sel(ws6, h, 4, str(g.get("bulanAsal") or ""))
    sel(ws6, h, 5, str(g.get("maTujuan") or ""))
    sel(ws6, h, 6, g.get("kapalTujuan") or "")
    sel(ws6, h, 7, str(g.get("jenisTujuan") or ""))
    sel(ws6, h, 8, g.get("nominal") or 0, RP, tebal=True)
    h += 1
sel(ws6, h, 2, "TOTAL", tebal=True, isi=ABU)
for c in range(3, 8): sel(ws6, h, c, "", isi=ABU)
sel(ws6, h, 8, geser_total, RP, tebal=True, isi=ABU)
ws6.freeze_panes = ws6.cell(awal6, 2)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — DATA REALISASI (rinci, tanpa vendor)
# ══════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("7. Data Realisasi")
ws7.sheet_view.showGridLines = False
r = judul(ws7, 2, "RINCIAN REALISASI PER DOKUMEN",
          "Dipakai bila direksi menanyakan asal sebuah angka. Nama penyedia tidak ditampilkan.", lebar=7)
h = kepala_tabel(ws7, r, ["", "Tanggal", "Jenis", "Nama pengadaan", "Mata Anggaran", "Nilai", "Sumber"],
                 [3, 12, 11, 62, 28, 18, 12])
awal7 = h
for d in sorted(dokumen, key=lambda d: (d["tanggal"], -d["nilai"])):
    sel(ws7, h, 2, d["tanggal"])
    sel(ws7, h, 3, d["jenis"].upper())
    sel(ws7, h, 4, d["nama"][:80])
    sel(ws7, h, 5, nama_ma(d["ma"]))
    sel(ws7, h, 6, d["nilai"], RP)
    sel(ws7, h, 7, "SPPBJ" if d["sumber"] == "sppbj" else "Non PR PO")
    h += 1
ws7.freeze_panes = ws7.cell(awal7, 2)
ws7.auto_filter.ref = f"B{awal7-1}:G{h-1}"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 8 — BAHAN SLIDE
# ══════════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("8. Bahan Slide")
ws8.sheet_view.showGridLines = False
r = judul(ws8, 2, "BAHAN SLIDE — TINGGAL DIPINDAH KE POWERPOINT",
          "Satu baris = satu slide. Kolom terakhir adalah kalimat untuk diucapkan, bukan untuk ditulis di slide.", lebar=6)
h = kepala_tabel(ws8, r, ["", "Slide", "Judul", "Isi (poin)", "Angka kunci", "Catatan pembicara"],
                 [3, 8, 34, 60, 26, 60])

def slide(no, jud, poin, angka, catatan):
    global h
    x = sel(ws8, h, 2, no, tebal=True, rata="center"); x.alignment = Alignment(horizontal="center", vertical="top")
    for kol, isi_sel, tebal in ((3, jud, True), (4, poin, False), (5, angka, True), (6, catatan, False)):
        c = sel(ws8, h, kol, isi_sel, tebal=tebal)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws8.row_dimensions[h].height = 62
    h += 1

bulan_over_teks = ", ".join(f"{BULAN_NAMA[int(b[5:7])-1]} {rl/pg*100:.0f}%" for b, pg, rl in BULAN_OVER) or "tidak ada"
slide(1, "Kendali Anggaran Teknik Ternate 2026",
      "• Pagu, realisasi, dan sisa\n• Tiga hal yang dimintakan keputusan",
      f"Pagu {TOT_PAGU/1e9:.2f} M · Realisasi {TOT_REAL/1e9:.2f} M",
      "Buka dengan angka besar dulu, jangan langsung ke rincian. Sebut bahwa seluruh angka berasal dari SPPBJ yang sama dengan yang dikirim ke pusat.")
slide(2, "Posisi hari ini",
      f"• Serapan {TOT_REAL/TOT_PAGU*100:.0f}% dari pagu\n• Sisa {TOT_PAGU-TOT_REAL:,.0f}\n• {len(BULAN_OVER)} bulan melewati pagu",
      f"Sisa {(TOT_PAGU-TOT_REAL)/1e6:,.0f} juta",
      f"Bulan yang melewati pagu: {bulan_over_teks}. Tekankan bahwa ini sudah terjadi, bukan rencana.")
slide(3, "Ke mana uangnya pergi",
      "• Susunan biaya menurut Mata Anggaran\n• Pos terbesar dan penyebabnya",
      (f"{nama_ma(max(real_ma, key=real_ma.get))} {max(real_ma.values())/1e6:,.0f} jt" if real_ma else "-"),
      "Pakai grafik batang bertumpuk di sheet 2. Jangan bacakan semua angka; tunjuk satu pos terbesar saja.")
if BOROS:
    x = BOROS[0]; pos = penyumbang(x["kapal"])
    slide(4, f"Kapal yang paling melampaui rencana: {x['kapal']}",
          "• Dibandingkan terhadap RKA-nya sendiri\n• Pos penyumbang terbesar\n• Usulan tindakan",
          f"{x['rasio']*100:.0f}% dari RKA",
          f"Realisasi {x['real']:,.0f} vs RKA {x['rka']:,.0f}. Penyumbang: "
          + "; ".join(f"{n} {v/1e6:,.0f} jt" for n, v, _ in pos) + ".")
slide(5, "Bukan sekadar mahal — bandingkan yang setara",
      "• Rp per GT\n• Umur kapal\n• Kapal yang sedang docking dipisah",
      "Lihat sheet 3",
      "Ini menjawab sanggahan 'kapal itu memang besar'. Tunjukkan kolom Rp per GT.")
slide(6, "Pergeseran anggaran",
      f"• {len(geser)} kali pergeseran\n• Total {geser_total:,.0f}\n• Menandakan rencana awal belum sesuai lapangan",
      f"{geser_total/1e6:,.0f} juta",
      "Pakai ini sebagai jembatan menuju usulan RKA 2027 yang lebih realistis.")
slide(7, "Proyeksi sampai Desember",
      f"• Rata-rata {rata_real/1e6:,.0f} juta per bulan\n• Perkiraan realisasi setahun {proyeksi_akhir/1e9:.2f} M",
      f"Perkiraan {proyeksi_akhir/1e9:.2f} M",
      "Sebut tegas ini proyeksi, bukan komitmen. Dasarnya rata-rata belanja bulan berjalan.")
slide(8, "Yang kami mintakan keputusan",
      "1. Tambahan pagu / realokasi\n2. Evaluasi teknis kapal berbiaya tinggi\n3. Dasar RKA 2027",
      f"Tambahan {KELEBIHAN/1e6:,.0f} juta",
      "Tutup dengan tiga permintaan ini. Setiap permintaan sudah ada angka dasarnya di sheet 1.")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 9 — CATATAN & ASUMSI
# ══════════════════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("9. Catatan & Asumsi")
ws9.sheet_view.showGridLines = False
r = judul(ws9, 2, "CATATAN, ASUMSI, DAN BATAS DATA",
          "Dibaca sebelum angka dipakai di forum resmi.", lebar=6)
ws9.column_dimensions["B"].width = 34
ws9.column_dimensions["C"].width = 96
catatan = [
    ("Sumber realisasi", "SPPBJ dan Non PR PO pada aplikasi Manajemen Report Teknik, tarikan "
     + datetime.date.today().strftime("%d %B %Y") + "."),
    ("Sumber pagu", "Persetujuan Pusat per bulan yang dicatat di aplikasi (menu Kendali Anggaran Rutin)."),
    ("Sumber RKA", "Berkas KONTROL ANGGARAN TERNATE.xlsx, folder Perencanaan Docking 2026."),
    ("Cara menghitung nilai", "Harga SPBJ dipakai bila dokumennya sudah memiliki harga final; bila belum, "
     "harga usulan. Nilai baris = harga x jumlah."),
    ("Barang persediaan", f"Pengadaan bertanda stok persediaan senilai {nilai_stok:,.0f} TIDAK dihitung "
     "sebagai penggerus pagu, mengikuti aturan yang dipakai aplikasi."),
    ("Satu baris beberapa kapal", "Nilainya dibagi rata ke kapal-kapal yang disebut, supaya total tetap sama."),
    ("Bulan docking pada RKA rutin", "Berkas cabang menyebar anggaran docking ke bulan pelaksanaannya. "
     "Bulan tersebut dikeluarkan dari RKA rutin agar perbandingan tidak melonjak."),
    ("Vendor", "Nama penyedia sengaja tidak ditampilkan di seluruh berkas ini."),
    ("Proyeksi", "Perkiraan memakai rata-rata belanja bulan berjalan. Bukan komitmen, dan tidak "
     "memperhitungkan rencana docking yang belum terbit SPPBJ-nya."),
    ("Bulan tanpa realisasi", "Agustus 2026 sudah punya pagu tetapi belum ada SPPBJ terbit saat data ditarik."),
]
h = kepala_tabel(ws9, r, ["", "Hal", "Keterangan"], [3, 34, 96])
for a, b in catatan:
    sel(ws9, h, 2, a, tebal=True)
    c = sel(ws9, h, 3, b); c.alignment = Alignment(wrap_text=True, vertical="top")
    ws9.row_dimensions[h].height = 30
    h += 1

# atur cetak
for sheet in wb.worksheets:
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_options.horizontalCentered = True

wb.save(KELUAR)
print("SELESAI:", KELUAR)
print(f"  bulan over: {[(b, f'{r/p*100:.0f}%') for b, p, r in BULAN_OVER]}")
print(f"  kelebihan: {KELEBIHAN:,.0f} | proyeksi setahun: {proyeksi_akhir:,.0f}")
print(f"  kapal terboros: {[(x['kapal'], f'{x['rasio']*100:.0f}%') for x in BOROS[:3]]}")
print(f"  sheet: {wb.sheetnames}")
